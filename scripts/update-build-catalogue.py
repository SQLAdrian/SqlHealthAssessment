#!/usr/bin/env python3
"""
SQL Server Build Catalogue Scraper — publish-time script.

Fetches SQL Server build data from Microsoft sources and produces
Config/sql-build-catalogue.json for use by the LiveMonitor app.

Sources:
  1. https://aka.ms/SQLServerbuilds → XLSX download (primary, redirects to direct file)
  2. Lifecycle product pages → support dates (scraped from XLSX "General info" sheet URLs)
  3. research_output/SQLVersionTable.csv → local fallback / cross-check

Output: Config/sql-build-catalogue.json
"""

import csv
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests

REQUIRED_VERSIONS = ["2008", "2008R2", "2012", "2014", "2016", "2017", "2019", "2022"]

VERSION_MAJOR_MAP = {
    "2008": 10, "2008R2": 10, "2012": 11, "2014": 12,
    "2016": 13, "2017": 14, "2019": 15, "2022": 16,
}

VERSION_MINOR_MAP = {
    "2008": 0, "2008R2": 50, "2012": 0, "2014": 0,
    "2016": 0, "2017": 0, "2019": 0, "2022": 0,
}

VERSION_TO_MAJOR_MINOR = {
    v: (VERSION_MAJOR_MAP[v], VERSION_MINOR_MAP[v]) for v in REQUIRED_VERSIONS
}

XLSX_URL = "https://aka.ms/SQLServerbuilds"
CSV_PATH = "research_output/SQLVersionTable.csv"
OUTPUT_PATH = "Config/sql-build-catalogue.json"

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent

# Sheet name → version key mapping (from the XLSX sheet names)
SHEET_TO_VERSION = {
    "2025": None,  # not in our 8, but present in XLSX
    "2022": "2022",
    "2019": "2019",
    "2017": "2017",
    "2016": "2016",
    "2014": "2014",
    "2012": "2012",
    "2008 R2": "2008R2",
    "2008": "2008",
}

LIFECYCLE_PAGE_CACHE = {}

def log(msg):
    print(f"[update-build-catalogue] {msg}", file=sys.stderr)


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def _http_get(url, timeout=30, headers=None):
    if headers is None:
        headers = {
            "User-Agent": "LiveMonitor-BuildCatalogue/1.0 (publish-time scraper)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
    return requests.get(url, headers=headers, allow_redirects=True, timeout=timeout)


# ---------------------------------------------------------------------------
# Source 1 — XLSX download + parse
# ---------------------------------------------------------------------------

def _normalize_date(date_raw):
    """Parse datetime or string to ISO date 'YYYY-MM-DD' or None."""
    if date_raw is None:
        return None
    if isinstance(date_raw, datetime):
        return date_raw.strftime("%Y-%m-%d")
    s = str(date_raw).strip()
    if not s or s.upper() in ("NULL", "NA", "N/A", ""):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def _classify_build_type(sp_level, cu_security_id, servicing_model):
    """Determine the build type from XLSX columns."""
    cu = str(cu_security_id).strip() if cu_security_id else ""
    svc = str(servicing_model).strip() if servicing_model else ""
    combined = f"{cu} {svc}".upper()

    if "GDR" in combined:
        return "GDR"
    if "CU" in combined or "CUMULATIVE" in combined:
        return "CU"
    if "RTM" in combined:
        return "RTM"
    if "SP" in combined or "SERVICE PACK" in combined:
        return "SP"
    if "AZURE" in combined or "CONNECT" in combined:
        return "Azure"
    if "SECURITY" in combined or "QFE" in combined:
        return "GDR"
    return "Other"


def _build_label(sp_level, cu_security_id):
    """Build a human-readable label from SP and CU/Security columns."""
    parts = []
    sp = str(sp_level).strip() if sp_level else ""
    cu = str(cu_security_id).strip() if cu_security_id else ""

    if sp and sp.upper() not in ("NA", "N/A", "NULL", ""):
        parts.append(sp)
    if cu and cu.upper() not in ("NA", "N/A", "NULL", ""):
        parts.append(cu)

    return " ".join(parts) if parts else cu if cu else ""


def _parse_xlsx_build_sheet(ws, version_str):
    """Parse a per-version sheet from the XLSX and return list of build dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Detect column indices
    col_idx = {}
    for i, h in enumerate(headers):
        if "build" in h and "number" in h:
            col_idx["build"] = i
        elif "kb number" in h or h == "kb":
            col_idx["kb"] = i
        elif "release date" in h:
            col_idx["release_date"] = i
        elif "service pack" in h and "level" in h:
            col_idx["sp"] = i
        elif "cumulative update" in h or "security id" in h:
            col_idx["cu"] = i
        elif "servicing model" in h:
            col_idx["servicing"] = i

    if "build" not in col_idx:
        return []

    major_ver, minor_ver = VERSION_TO_MAJOR_MINOR[version_str]
    prefix = f"{major_ver}.{minor_ver}."

    builds = []
    seen = set()

    for row in rows[1:]:
        build_raw = row[col_idx["build"]] if col_idx["build"] < len(row) else None
        if build_raw is None:
            continue

        build_str = str(build_raw).strip()
        if not build_str:
            continue

        # Normalize build number
        build_str = re.sub(r'\s+', '', build_str)

        # Verify this build belongs to expected version
        if not build_str.startswith(prefix):
            # Check if the major version part matches at least
            parts = build_str.split(".")
            if len(parts) >= 2:
                try:
                    mv = int(parts[0])
                    minv = int(parts[1])
                    if mv != major_ver or minv != minor_ver:
                        continue
                except ValueError:
                    continue

        if build_str in seen:
            continue
        seen.add(build_str)

        # KB number
        kb_raw = row[col_idx["kb"]] if "kb" in col_idx and col_idx["kb"] < len(row) else None
        kb = None
        if kb_raw is not None:
            kb_str_raw = str(kb_raw).strip()
            if kb_str_raw and kb_str_raw.upper() not in ("NA", "N/A", "NULL", "", "NONE"):
                try:
                    kb_num = int(kb_raw)
                    kb = f"KB{kb_num}"
                except (ValueError, TypeError):
                    if kb_str_raw.upper().startswith("KB"):
                        kb = kb_str_raw
                    else:
                        kb = f"KB{kb_str_raw}"

        # Release date
        date_raw = row[col_idx["release_date"]] if "release_date" in col_idx and col_idx["release_date"] < len(row) else None
        release_date = _normalize_date(date_raw)

        # SP level
        sp_raw = row[col_idx["sp"]] if "sp" in col_idx and col_idx["sp"] < len(row) else None

        # CU/Security ID
        cu_raw = row[col_idx["cu"]] if "cu" in col_idx and col_idx["cu"] < len(row) else None

        # Servicing model
        svc_raw = row[col_idx["servicing"]] if "servicing" in col_idx and col_idx["servicing"] < len(row) else None

        label = _build_label(sp_raw, cu_raw)
        btype = _classify_build_type(sp_raw, cu_raw, svc_raw)

        entry = {
            "build": build_str,
            "label": label,
            "releaseDate": release_date,
            "type": btype,
        }
        if kb:
            entry["kb"] = kb

        builds.append(entry)

    return builds


def fetch_xlsx_builds():
    """Download and parse the XLSX. Returns (builds_by_version, lifecycle_urls) or (None, None) on failure."""
    log("Downloading XLSX from aka.ms/SQLServerbuilds ...")

    try:
        resp = _http_get(XLSX_URL, timeout=60)
        resp.raise_for_status()
    except Exception as e:
        log(f"ERROR: XLSX download failed: {e}")
        return None, None

    content_type = resp.headers.get("Content-Type", "")
    if "html" in content_type.lower() or "text" in content_type.lower():
        log("WARNING: XLSX redirect gave HTML, not binary — attempting to find download link")
        m = re.search(r'href="(https?://[^"]+\.xlsx)"', resp.text, re.IGNORECASE)
        if m:
            xlsx_url = m.group(1)
            log(f"  Found XLSX link: {xlsx_url}")
            try:
                resp = _http_get(xlsx_url, timeout=60)
                resp.raise_for_status()
            except Exception as e:
                log(f"ERROR: XLSX secondary download failed: {e}")
                return None, None
        else:
            log("ERROR: Could not find XLSX download URL in HTML page")
            return None, None

    tmp_path = REPO_ROOT / "scripts" / "_temp_builds.xlsx"
    tmp_path.write_bytes(resp.content)
    log(f"  Downloaded {len(resp.content)} bytes")

    try:
        wb = openpyxl.load_workbook(str(tmp_path), read_only=True, data_only=True)

        # Parse "General info" sheet for lifecycle URLs
        lifecycle_urls = {}
        if "General info" in wb.sheetnames:
            ws_info = wb["General info"]
            info_rows = list(ws_info.iter_rows(values_only=True))
            for row in info_rows[1:]:
                version_name = str(row[0]).strip() if row[0] else ""
                url = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if not url:
                    continue
                # Map "SQL Server 2022" → "2022"
                m = re.search(r'SQL\s*Server\s*(\d{4})\s*(R2)?', version_name, re.IGNORECASE)
                if m:
                    ver = m.group(1)
                    if m.group(2):
                        ver = ver + "R2"
                    if ver in REQUIRED_VERSIONS:
                        lifecycle_urls[ver] = url

        # Parse per-version sheets
        builds_by_version = {}
        for sheet_name, version_key in SHEET_TO_VERSION.items():
            if version_key is None or version_key not in REQUIRED_VERSIONS:
                continue
            if sheet_name not in wb.sheetnames:
                log(f"  WARNING: sheet '{sheet_name}' not found in XLSX")
                continue

            ws = wb[sheet_name]
            ver_builds = _parse_xlsx_build_sheet(ws, version_key)
            if ver_builds:
                # Sort by build number
                def _sort_key(b):
                    parts = b["build"].split(".")
                    return tuple(int(p) if p.isdigit() else 0 for p in parts)
                ver_builds.sort(key=_sort_key)
                builds_by_version[version_key] = ver_builds
                log(f"  {version_key}: {len(ver_builds)} builds from XLSX")

        wb.close()
    except Exception as e:
        log(f"ERROR: XLSX parse failed: {e}")
        return None, None
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    return builds_by_version, lifecycle_urls


# ---------------------------------------------------------------------------
# Source 2 — Lifecycle dates (scraped from individual product pages)
# ---------------------------------------------------------------------------

def _scrape_lifecycle_page(url):
    """Scrape a single lifecycle product page for support dates."""
    if url in LIFECYCLE_PAGE_CACHE:
        return LIFECYCLE_PAGE_CACHE[url]

    try:
        resp = _http_get(url, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        log(f"  WARNING: lifecycle page {url} failed: {e}")
        LIFECYCLE_PAGE_CACHE[url] = (None, None)
        return None, None

    html = resp.text

    # Find all <local-time> datetime attributes in order
    # The support dates table has columns: Start Date, Mainstream End Date, Extended End Date
    datetimes = re.findall(r'<local-time[^>]+format="date"[^>]+datetime="([^"]+)"', html)
    if not datetimes:
        datetimes = re.findall(r'datetime="(\d{4}-\d{2}-\d{2}[^"]*)"', html)

    # Extract dates (just YYYY-MM-DD part)
    dates = []
    for dt in datetimes:
        m = re.match(r'(\d{4}-\d{2}-\d{2})', dt)
        if m:
            dates.append(m.group(1))

    mainstream = None
    extended = None

    # Typical structure: dates[0]=start, dates[1]=mainstream end, dates[2]=extended end
    if len(dates) >= 2:
        mainstream = dates[1]
    if len(dates) >= 3:
        extended = dates[2]

    LIFECYCLE_PAGE_CACHE[url] = (mainstream, extended)
    return mainstream, extended


def fetch_lifecycle_dates(lifecycle_urls=None):
    """Get lifecycle support dates for all required versions."""
    log("Fetching lifecycle support dates ...")

    result = {}
    for ver in REQUIRED_VERSIONS:
        mainstream = None
        extended = None

        if lifecycle_urls and ver in lifecycle_urls:
            mainstream, extended = _scrape_lifecycle_page(lifecycle_urls[ver])
        else:
            # Try standard URL pattern
            url_map = {
                "2008": "https://learn.microsoft.com/lifecycle/products/microsoft-sql-server-2008",
                "2008R2": "https://learn.microsoft.com/lifecycle/products/microsoft-sql-server-2008-r2",
                "2012": "https://learn.microsoft.com/lifecycle/products/microsoft-sql-server-2012",
                "2014": "https://learn.microsoft.com/lifecycle/products/sql-server-2014",
                "2016": "https://learn.microsoft.com/lifecycle/products/sql-server-2016",
                "2017": "https://learn.microsoft.com/lifecycle/products/sql-server-2017",
                "2019": "https://learn.microsoft.com/lifecycle/products/sql-server-2019",
                "2022": "https://learn.microsoft.com/lifecycle/products/sql-server-2022",
            }
            if ver in url_map:
                mainstream, extended = _scrape_lifecycle_page(url_map[ver])

        if mainstream or extended:
            log(f"  {ver}: mainstream={mainstream}, extended={extended}")
        else:
            log(f"  {ver}: no lifecycle dates found")

        result[ver] = {
            "mainstreamSupportEnds": mainstream,
            "extendedSupportEnds": extended,
        }

    return result


# ---------------------------------------------------------------------------
# Source 3 — CSV fallback / cross-check
# ---------------------------------------------------------------------------

def parse_csv_fallback():
    """Parse research_output/SQLVersionTable.csv → dict version→[builds]."""
    csv_path = REPO_ROOT / CSV_PATH
    if not csv_path.exists():
        log(f"ERROR: CSV fallback not found at {csv_path}")
        return None

    log(f"Parsing CSV fallback: {csv_path}")
    result = {}
    with open(csv_path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            server = row.get("Server", "").strip()
            if server not in REQUIRED_VERSIONS:
                continue

            build_raw = row.get("Build", "").strip()
            if not build_raw:
                continue

            build_str = re.sub(r'\s+', '', build_raw)
            if not build_str:
                continue

            sp = row.get("ServicePack", "").strip()
            cu = row.get("CumulativeUpdate", "").strip()
            release = row.get("ReleaseDate", "").strip()

            label = _build_label(sp, cu)
            btype = _classify_build_type(sp, cu, "")
            date = _normalize_date(release) if release else None

            entry = {
                "build": build_str,
                "label": label,
                "releaseDate": date,
                "type": btype,
            }

            result.setdefault(server, []).append(entry)

    # Dedup and sort within each version
    for ver in result:
        seen = set()
        unique = []
        for b in result[ver]:
            if b["build"] not in seen:
                seen.add(b["build"])
                unique.append(b)
        def _sort_key(b):
            parts = b["build"].split(".")
            return tuple(int(p) if p.isdigit() else 0 for p in parts)
        unique.sort(key=_sort_key)
        result[ver] = unique

    return result


# ---------------------------------------------------------------------------
# Cross-check
# ---------------------------------------------------------------------------

def cross_check(xlsx_builds, csv_builds):
    """Compare XLSX results against CSV. Flag release date disagreements >7 days."""
    disagreements = []
    for ver in REQUIRED_VERSIONS:
        xl = xlsx_builds.get(ver, [])
        cs = csv_builds.get(ver, [])
        if not xl or not cs:
            continue

        cs_by_build = {b["build"]: b for b in cs if b.get("releaseDate")}

        for xb in xl:
            xb_date = xb.get("releaseDate")
            cb = cs_by_build.get(xb["build"])
            if not cb or not xb_date:
                continue
            cb_date = cb.get("releaseDate")
            if not cb_date:
                continue

            try:
                d1 = datetime.strptime(xb_date, "%Y-%m-%d")
                d2 = datetime.strptime(cb_date, "%Y-%m-%d")
                delta = abs((d1 - d2).days)
                if delta > 7:
                    disagreements.append({
                        "version": ver,
                        "build": xb["build"],
                        "xlsxDate": xb_date,
                        "csvDate": cb_date,
                        "deltaDays": delta,
                    })
            except ValueError:
                continue

    if disagreements:
        log(f"CROSS-CHECK: {len(disagreements)} date disagreements >7 days:")
        for d in disagreements:
            log(f"  {d['version']} / {d['build']} — XLSX:{d['xlsxDate']} vs CSV:{d['csvDate']} ({d['deltaDays']}d)")
    else:
        log("CROSS-CHECK: No significant date disagreements found.")

    return disagreements


# ---------------------------------------------------------------------------
# Output builder
# ---------------------------------------------------------------------------

def build_output(builds_by_version, lifecycle_dates, sources_used, cross_check_issues):
    """Assemble the final output dict."""
    output = {
        "lastUpdated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "sources": sources_used,
        "versions": {},
    }

    for ver in REQUIRED_VERSIONS:
        builds = builds_by_version.get(ver, [])
        latest = builds[-1] if builds else None
        latest_build = latest["build"] if latest else None
        latest_release_date = latest.get("releaseDate") if latest else None

        lc = lifecycle_dates.get(ver, {})
        mainstream = lc.get("mainstreamSupportEnds")
        extended = lc.get("extendedSupportEnds")

        output["versions"][ver] = {
            "majorBuild": VERSION_MAJOR_MAP[ver],
            "mainstreamSupportEnds": mainstream,
            "extendedSupportEnds": extended,
            "latestBuild": latest_build,
            "latestReleaseDate": latest_release_date,
            "builds": builds,
        }

    if cross_check_issues:
        output["crossCheckIssues"] = cross_check_issues

    return output


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_output(output):
    """Check acceptance gates. Returns (bool, list of error messages)."""
    errors = []

    versions_present = list(output["versions"].keys())
    missing = [v for v in REQUIRED_VERSIONS if v not in versions_present]
    if missing:
        errors.append(f"Missing versions in output: {missing}")

    for ver in REQUIRED_VERSIONS:
        vdata = output["versions"].get(ver)
        if not vdata:
            errors.append(f"Version '{ver}' has no data at all")
            continue

        build_count = len(vdata.get("builds", []))
        if build_count == 0:
            errors.append(f"Version '{ver}' has ZERO builds — unacceptable")
        elif build_count < 5:
            errors.append(f"Version '{ver}' has only {build_count} build(s) — minimum 5 expected")

        if vdata.get("latestBuild") is None:
            errors.append(f"Version '{ver}' has no latestBuild")

    if errors:
        return False, errors
    return True, []


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.chdir(str(REPO_ROOT))

    sources_used = {
        "primary": XLSX_URL,
        "lifecycle": "Lifecycle product pages (scraped)",
        "fallback": CSV_PATH,
    }
    cross_check_issues = []

    # 1. Fetch XLSX (includes lifecycle URLs)
    xlsx_builds, lifecycle_urls = fetch_xlsx_builds()

    # 2. Fetch lifecycle dates
    lifecycle_dates = fetch_lifecycle_dates(lifecycle_urls)

    # 3. Determine primary source
    if xlsx_builds and all(ver in xlsx_builds and len(xlsx_builds.get(ver, [])) > 0 for ver in REQUIRED_VERSIONS):
        log("XLSX fetch succeeded with all versions present.")
        builds_by_version = xlsx_builds

        # Cross-check against CSV
        csv_result = parse_csv_fallback()
        if csv_result:
            cross_check_issues = cross_check(xlsx_builds, csv_result)
    else:
        if xlsx_builds:
            missing_vers = [v for v in REQUIRED_VERSIONS if v not in xlsx_builds or len(xlsx_builds.get(v, [])) == 0]
            log(f"XLSX missing versions: {missing_vers} — falling back to CSV")
        else:
            log("XLSX fetch failed — falling back to CSV")

        csv_result = parse_csv_fallback()
        if not csv_result:
            log("FATAL: Both XLSX and CSV fallback failed. Cannot produce catalogue.")
            sys.exit(1)

        builds_by_version = csv_result
        sources_used["primary"] = f"fallback: {CSV_PATH}"
        log(f"CSV fallback loaded {sum(len(v) for v in csv_result.values())} builds across {len(csv_result)} versions.")

    # 4. Build output
    output = build_output(builds_by_version, lifecycle_dates, sources_used, cross_check_issues)

    # 5. Validate
    ok, errors = validate_output(output)
    if not ok:
        for e in errors:
            log(f"VALIDATION ERROR: {e}")
        log("FATAL: Validation failed — exiting with non-zero code.")
        sys.exit(1)

    # 6. Write output
    output_path = REPO_ROOT / OUTPUT_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # 7. Summary
    total_builds = sum(len(v["builds"]) for v in output["versions"].values())
    log(f"SUCCESS: Wrote {output_path}")
    log(f"  Versions: {len(output['versions'])}")
    log(f"  Total builds: {total_builds}")
    for ver in REQUIRED_VERSIONS:
        vd = output["versions"][ver]
        log(f"  {ver}: {len(vd['builds'])} builds, latest={vd['latestBuild']}")


if __name__ == "__main__":
    main()
